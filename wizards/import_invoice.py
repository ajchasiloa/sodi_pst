import base64
import binascii
import csv
import datetime
import io
import tempfile
import openpyxl  
from odoo.exceptions import ValidationError
from odoo import fields, models


class ImportInvoice(models.TransientModel):
    """Model for import invoice"""
    _name = 'import.invoice'
    _description = 'Invoice Import'

    file_type = fields.Selection(
        selection=[('csv', 'CSV File'), ('xlsx', 'XLSX File')],
        string='Import File Type', default='csv',
        help="It helps to choose the file type"
    )
    file = fields.Binary(string="File", help="File")
    update_posted = fields.Boolean(
        string='Update Posted Record?',
        help='If enabled, the records in "Posted" state will be converted to draft'
             ' and values are updated. These records will then again be posted'
             ' if "Post Automatically" is activated'
    )
    auto_post = fields.Boolean(string='Post Automatically',
                               help="Post Automatically"
                               )
    journal = fields.Selection(
        selection=[('Bank', 'Bank'), ('Cash', 'Cash')],
        string='Journal', default='Bank', help='It helps to choose Journal type'
    )
    order_number = fields.Selection(
        selection=[('from_system', 'From System'), ('from_file', 'From File')],
        string='Number', default='from_file', help="Order number"
    )
    import_product_by = fields.Selection(
        selection=[('name', 'Name'), ('default_code', 'Internal Reference'),
                   ('barcode', 'Barcode')], required=True, default="name",
        string="Import invoice by", help="Product import"
    )
    type = fields.Selection(
        selection=[('out_invoice', 'Invoice'), ('in_invoice', 'Bill'),
                   ('out_refund', 'Credit Note'), ('in_refund', 'Refund')],
        string='Invoicing Type', required=True, help="Invoice type",
        default="out_invoice"
    )

    def action_import_invoice(self):
        """Importar facturas usando encabezados en español"""

        account_move = self.env['account.move']
        account_account = self.env['account.account']
        uom_uom = self.env['uom.uom']

        items = self.read_file()

        # Validar campos requeridos
        error_msg = ""
        for row_index, item in enumerate(items, start=2):
            if not item or len(item) < 6:
                error_msg += f"Fila {row_index} está vacía o incompleta.\n"
                continue

            required_fields = ['Socio', 'Fecha de Factura', 'Producto', 'Cantidad', 'Precio', 'Número']
            for field in required_fields:
                if not item.get(field):
                    error_msg += f"El campo '{field}' falta en la fila {row_index}\n"

        if error_msg:
            raise ValidationError(f"Error de validación:\n{error_msg}")

        imported = 0
        confirmed = 0
        imported_invoices = []

        for row_index, item in enumerate(items, start=2):
            # Primero validamos si el cliente existe
            partner_name = item.get('Socio')
            partner = self.validar_cliente(partner_name)  # Llamada de validación al cliente
            
            # Validación de producto
            product_name = item.get('Producto')
            product = self.validar_producto(product_name)

            # Validación de fecha de factura
            invoice_date = self.parsear_fecha(item['Fecha de Factura'], row_index)

            # Validación de cantidad y precio
            try:
                cantidad = float(item['Cantidad'])
                precio = float(item['Precio'])
            except ValueError:
                raise ValidationError(f"Cantidad o precio inválido en la fila {row_index}")

            # Validación de número de factura único
            invoice_number = item.get('Número')
            existing_invoice = account_move.search([
                ('name', '=', invoice_number),
                ('move_type', '=', self.type)
            ], limit=1)
            
            if existing_invoice:
                raise ValidationError(f"⚠ La factura con número '{invoice_number}' ya está registrada en el sistema. No se creará esta factura.")

            # Preparar valores para la factura
            vals = {
                'move_type': self.type,
                'partner_id': partner.id,
                'invoice_date': invoice_date,
                'name': invoice_number
            }

            # Fecha de vencimiento (opcional)
            if item.get('Fecha de Vencimiento'):
                vals['invoice_date_due'] = self.parsear_fecha(
                    item['Fecha de Vencimiento'], 
                    row_index, 
                    campo='Fecha de Vencimiento'
                )

            # Crear factura
            invoice = account_move.create(vals)

            # Preparar línea de factura
            line_vals = {
                'product_id': product.id,
                'quantity': cantidad,
                'price_unit': precio
            }

            # Cuenta contable (opcional)
            if item.get('Código de Cuenta'):
                account = account_account.search([('code', '=', str(item['Código de Cuenta']).strip())], limit=1)
                if account:
                    line_vals['account_id'] = account.id

            # Unidad de medida (opcional)
            if item.get('UoM'):
                uom = uom_uom.search([('name', '=', item['UoM'])], limit=1)
                if uom:
                    line_vals['product_uom_id'] = uom.id

            # Agregar línea de factura
            invoice.write({'invoice_line_ids': [(0, 0, line_vals)]})
            
            imported += 1
            imported_invoices.append(invoice)

        # Post if needed
        if self.auto_post and imported_invoices:
            for inv in set(imported_invoices):
                inv.action_post()
                confirmed += 1

        # Retorno de mensaje de éxito
        return {
            'effect': {
                'fadeout': 'slow',
                'message': f"Imported {imported} records.\nPosted {confirmed} records"
            }
        }

    def action_test_import_invoice(self):
        """Test import file without importing data."""
        if not self.file:
            raise ValidationError("Por favor, sube un archivo válido antes de continuar.")

        # Validación del tipo de archivo
        if self.file_type not in ['csv', 'xlsx']:
            raise ValidationError("Invalid file type. Only CSV and XLSX are allowed.")

        try:
            if self.file_type == 'csv':
                # Verificar que el archivo CSV tenga contenido
                csv_data = base64.b64decode(self.file)
                data_file = io.StringIO(csv_data.decode("utf-8"))
                data_file.seek(0)
                reader = csv.reader(data_file)
                rows = list(reader)
                if len(rows) <= 1:
                    raise ValidationError("The CSV file is empty or missing data.")
            elif self.file_type == 'xlsx':
                # Verificar que el archivo XLSX tenga contenido
                fp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
                fp.write(binascii.a2b_base64(self.file))
                fp.seek(0)
                workbook = openpyxl.load_workbook(fp.name)
                sheet = workbook.active
                if sheet.max_row <= 1:  # Si solo tiene una fila (cabecera), es un archivo vacío
                    raise ValidationError("The XLSX file is empty or missing data.")
        except Exception as e:
            raise ValidationError(f"Error reading the file: {str(e)}")

        # Si las validaciones son exitosas
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': 'Validation Success',
                'message': 'The file was validated successfully.',
                'sticky': False,
            }
        }

    def read_file(self):
        """Method to read the file depending on its type"""
        if self.file_type == 'csv':
            return self.read_csv_file()
        elif self.file_type == 'xlsx':
            return self.read_xlsx_file()

    def read_csv_file(self):
        """Leer un archivo CSV"""
        try:
            csv_data = base64.b64decode(self.file)
            data_file = io.StringIO(csv_data.decode("utf-8"))
            data_file.seek(0)
            csv_reader = csv.DictReader(data_file, delimiter=',')
            return list(csv_reader)
        except Exception as e:
            raise ValidationError(f"Archivo CSV no válido. Error: {e}")

    def read_xlsx_file(self):
        """Read an XLSX file"""
        try:
            fp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            fp.write(binascii.a2b_base64(self.file))
            fp.seek(0)
            workbook = openpyxl.load_workbook(fp.name)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]
            rows = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
            return rows
        except Exception as e:
            raise ValidationError(f"Archivo XLSX no válido. Error: {e}")

    def action_generate_template(self):
        """Genera una plantilla Excel para importar facturas"""

        import base64
        from io import BytesIO
        import xlsxwriter

        # Definimos los encabezados requeridos
        field_labels = [
            "Socio",                   # Partner
            "Fecha de Factura",        # Invoice Date
            "Fecha de Vencimiento",    # Due Date
            "Número",                  # Number
            "Producto",                # Product
            "Código de Cuenta",        # Account Code
            "UoM",                     # Uom
            "Cantidad",                # Quantity
            "Precio"                   # Price
        ]

        # Crear archivo Excel en memoria
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet("Invoice Import Template")

        # Estilo para encabezado
        header_format = workbook.add_format({'bold': True, 'bg_color': '#D9D9D9', 'align': 'center', 'valign': 'vcenter'})
        
        # Escribir encabezados en la primera fila
        for col, label in enumerate(field_labels):
            worksheet.write(0, col, label, header_format)

        # Añadir una línea de ejemplo
        example_data = [
            "DALMART",                # Socio
            "2024-03-19 08:00:00",        # Fecha de Factura
            "2024-04-19 08:00:00",        # Fecha de Vencimiento
            "INV-12345",                  # Número
            "Producto A",                 # Producto
            "12345",                       # Código de Cuenta
            "Unidad",                     # UoM
            10.0,                          # Cantidad
            100.0                          # Precio
        ]

        # Escribir los datos de ejemplo en la segunda fila
        for col, value in enumerate(example_data):
            worksheet.write(1, col, value)

        # Ajustar automáticamente el tamaño de las columnas para que todo el texto sea visible
        for col in range(len(field_labels)):
            # Ajusta el ancho de las columnas al tamaño máximo entre el encabezado y los datos
            column_width = max(len(field_labels[col]), max(len(str(example_data[col])) for example_data in [example_data]))
            worksheet.set_column(col, col, column_width)

        # Cerrar y preparar el archivo
        workbook.close()
        output.seek(0)

        # Crear adjunto en Odoo
        attachment = self.env['ir.attachment'].create({
            'name': 'plantilla_importación_facturas.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(output.read()).decode(),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        # Retornar archivo para descarga
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    def parsear_fecha(self, date_str, row_index, campo='Fecha de Factura'):
        """
        Parsea una fecha de manera flexible, manejando múltiples formatos.
        
        :param date_str: Cadena de fecha a parsear
        :param row_index: Índice de la fila para mensajes de error
        :param campo: Nombre del campo de fecha para mensajes de error
        :return: Objeto de fecha parseado
        """
        if not date_str:
            raise ValidationError(f"El campo '{campo}' está vacío en la fila {row_index}")
        
        # Convertir a cadena y eliminar espacios
        date_str = str(date_str).strip()
        
        # Lista de formatos de fecha a intentar
        formatos_fecha = [
            '%Y-%m-%d',           # YYYY-MM-DD
            '%d/%m/%Y',           # DD/MM/YYYY
            '%m/%d/%Y',           # MM/DD/YYYY
            '%Y/%m/%d',           # YYYY/MM/DD
            '%d-%m-%Y',           # DD-MM-YYYY
            '%Y.%m.%d',           # YYYY.MM.DD
        ]
        
        # Intentar parsear con diferentes formatos
        for formato in formatos_fecha:
            try:
                # Si hay un espacio, tomar solo la primera parte (fecha)
                if ' ' in date_str:
                    date_str = date_str.split()[0]
                
                fecha_parseada = datetime.datetime.strptime(date_str, formato).date()
                return fecha_parseada
            except ValueError:
                continue
        
        # Si ningún formato funciona, lanzar error
        raise ValidationError(f"Formato de {campo} inválido en la fila {row_index}. Formatos aceptados: YYYY-MM-DD, DD/MM/YYYY, etc.")

    def validar_cliente(self, partner_name):
        """
        Busca el cliente en el sistema de Odoo por su nombre exacto o similar.
        Si el cliente no existe o no se encuentra una coincidencia, lanza un error de validación.
        
        :param partner_name: Nombre exacto o similar del partner a buscar
        :return: Registro del partner encontrado
        """
        # Buscar un partner exacto
        partner = self.env['res.partner'].search([('name', '=', partner_name)], limit=1)
        
        if not partner:
            # Si no se encuentra el partner exacto, buscar por nombre similar
            partner = self.env['res.partner'].search([('name', 'ilike', partner_name)], limit=1)
        
        if not partner:
            # Si no se encuentra ni exacto ni similar, arrojar un error
            raise ValidationError(f"Usuario no encontrado: El socio '{partner_name}' no está registrado en el sistema.")
        
        return partner

    def validar_producto(self, product_name):
        """
        Busca el producto en el sistema de Odoo por su nombre exacto.
        Si el producto no existe, lanza un error de validación.
        
        :param product_name: Nombre exacto del producto a buscar
        :return: Registro del producto encontrado
        """
        product = self.env['product.product'].search([('name', '=', product_name)], limit=1)
        
        if not product:
            raise ValidationError(f"El producto '{product_name}' no está registrado en el sistema. No se puede procesar la factura.")
        
        return product